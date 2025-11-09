import os
import json
import sys

IMAGE_EXTS = {'.png', '.jpg', '.jpeg', '.webp', '.gif', '.bmp'}


def is_image(filename):
    return os.path.splitext(filename)[1].lower() in IMAGE_EXTS


def read_metadata_from_excel(base_dir, excel_name='dataSet.xlsx'):
    """Read metadata from the Excel file and return two mappings inside a dict:

    {
      'by_key': { exact_key: meta_dict },           # exact_key is either Item or 'ID Item'
      'by_item': { item_name: [meta_dict, ...] }    # list to detect duplicates
    }

    Each meta_dict contains fields from desired_cols (Collection/Colection, Type, Price PLN, Size (EU), Fabric, Base, ID).
    """
    excel_path = os.path.join(base_dir, excel_name)
    if not os.path.isfile(excel_path):
        print(f"Excel file not found at {excel_path}; continuing without metadata.")
        return {}

    # Accept both 'collection' and common misspelling 'colection' in the Excel headers
    desired_cols = {
        'state': 'State',
        'colection': 'Colection',
        'type': 'Type',
        'price pln': 'Price PLN',
        'size (eu)': 'Size (EU)',
        'fabric': 'Fabric',
        'base': 'Base',
        'id': 'ID',
        'item': 'Item',
        'website des': 'Website Des',
        


    }

    def _make_empty_meta():
        return {v: None for k, v in desired_cols.items() if v != 'Item'}

    def _normalize_id_value(val):
        if val is None:
            return None
        # try to convert numeric-looking values to int when possible
        try:
            f = float(val)
        except Exception:
            return str(val).strip()
        if f.is_integer():
            return int(f)
        return f

    # Try pandas first (convenient). If not available, fall back to openpyxl.
    try:
        import pandas as pd
        df = pd.read_excel(excel_path, engine='openpyxl' if 'openpyxl' in sys.modules else None)
        # build header lookup (lowercase -> actual column name)
        cols = {c.lower(): c for c in df.columns}
        if 'item' not in cols:
            print('Excel found but required column "Item" not present; continuing without metadata.')
            return {}
        print(f"Excel loaded: columns found: {list(df.columns)}")
        print(f"Total rows in Excel (including header): {len(df)}")
        by_key = {}
        by_item = {}
        for _, row in df.iterrows():
            item_val = row[cols['item']]
            if pd.isna(item_val):
                continue
            item_name = str(item_val).strip()
            meta = _make_empty_meta()
            # fill available fields
            for low_col, pretty in desired_cols.items():
                if low_col == 'item':
                    continue
                if low_col in cols:
                    val = row[cols[low_col]]
                    meta[pretty] = None if pd.isna(val) else (str(val).strip() if not isinstance(val, (int, float)) else val)

            # append to by_item list
            by_item.setdefault(item_name, []).append(meta)
            # exact key for item_name
            by_key[item_name] = meta
            # composite key if ID present
            id_val = meta.get('ID')
            if id_val is not None:
                # normalize id string for composite key
                id_norm = str(id_val) if not isinstance(id_val, (int, float)) else (str(int(id_val)) if float(id_val).is_integer() else str(id_val))
                composite = f"{id_norm} {item_name}"
                by_key[composite] = meta

        print(f"Metadata entries parsed: {sum(len(v) for v in by_item.values())}")
        # detect duplicates
        dup_count = sum(1 for k, v in by_item.items() if len(v) > 1)
        if dup_count:
            print(f"Warning: {dup_count} Item names have multiple metadata rows (duplicates). Prefer folder names with ID prefix to disambiguate.")
        sample_keys = list(by_key.keys())[:10]
        if sample_keys:
            print(f"Sample keys from Excel mapping (exact/composite): {sample_keys}")
        return {'by_key': by_key, 'by_item': by_item}
    except Exception:
        # fallback to openpyxl without pandas
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, read_only=True)
            ws = wb.active
            header_cells = next(ws.iter_rows(min_row=1, max_row=1))
            headers = [str(c.value).strip().lower() if c.value is not None else '' for c in header_cells]
            if 'item' not in headers:
                print('Excel found but required column "Item" not present; continuing without metadata.')
                return {}
            idx_map = {h: i for i, h in enumerate(headers)}
            by_key = {}
            by_item = {}
            for row in ws.iter_rows(min_row=2):
                item_cell = row[idx_map['item']].value
                if item_cell is None:
                    continue
                item_name = str(item_cell).strip()
                meta = _make_empty_meta()
                for low_col, pretty in desired_cols.items():
                    if low_col == 'item':
                        continue
                    if low_col in idx_map:
                        cell = row[idx_map[low_col]].value
                        meta[pretty] = None if cell is None else (str(cell).strip() if not isinstance(cell, (int, float)) else cell)

                by_item.setdefault(item_name, []).append(meta)
                by_key[item_name] = meta
                id_val = meta.get('ID')
                if id_val is not None:
                    id_norm = str(id_val) if not isinstance(id_val, (int, float)) else (str(int(id_val)) if float(id_val).is_integer() else str(id_val))
                    composite = f"{id_norm} {item_name}"
                    by_key[composite] = meta

            print(f"Excel loaded via openpyxl: headers found: {headers}")
            print(f"Metadata entries parsed: {sum(len(v) for v in by_item.values())}")
            dup_count = sum(1 for k, v in by_item.items() if len(v) > 1)
            if dup_count:
                print(f"Warning: {dup_count} Item names have multiple metadata rows (duplicates). Prefer folder names with ID prefix to disambiguate.")
            sample_keys = list(by_key.keys())[:10]
            if sample_keys:
                print(f"Sample keys from Excel mapping (exact/composite): {sample_keys}")
            return {'by_key': by_key, 'by_item': by_item}
        except Exception as e:
            print('Failed to read Excel metadata; please install pandas and openpyxl or ensure the file is a valid xlsx. Error:', e)
            return {}


def gather_images(base_dir, metadata_map):
    """Gather images from Items/*/Website and return a dict in the shape { 'Items': { item_name: entry } }.

    This preserves the previous top-level category mapping so downstream consumers that iterate
    over index.values() continue to work.
    """
    result = {'Items': {}}
    items_dir = os.path.join(base_dir, 'Items')

    # metadata_map may be the new structure {'by_key':..., 'by_item':...} or the legacy flat mapping
    if metadata_map and isinstance(metadata_map, dict) and 'by_key' in metadata_map:
        by_key = metadata_map.get('by_key', {})
        by_item = metadata_map.get('by_item', {})
    else:
        by_key = metadata_map or {}
        by_item = {}

    if not os.path.isdir(items_dir):
        return result

    for item_name in sorted(os.listdir(items_dir)):
        item_path = os.path.join(items_dir, item_name)
        if not os.path.isdir(item_path):
            continue
        website_dir = os.path.join(item_path, 'Website')
        if not os.path.isdir(website_dir):
            continue
        images = []
        for fname in sorted(os.listdir(website_dir)):
            fpath = os.path.join(website_dir, fname)
            if os.path.isfile(fpath) and is_image(fname):
                # store path relative to base_dir (dataSet)
                rel = os.path.relpath(fpath, base_dir).replace('\\', '/')
                images.append(rel)
        if images:
            # If folder name is like "10 Dusty Plate Pants", split leading ID and item
            id_from_folder = None
            item_key = item_name
            parts = item_name.split(' ', 1)
            if len(parts) == 2 and parts[0].isdigit():
                id_from_folder = parts[0]
                item_key = parts[1]

            # Try matching metadata by exact folder name first, then by item_key
            meta = None
            # exact/composite key match
            if item_name in by_key:
                meta = by_key.get(item_name)
            elif item_key in by_key:
                meta = by_key.get(item_key)
            else:
                # fallback: multiple metadata rows for this item name
                candidates = by_item.get(item_key)
                if candidates:
                    if len(candidates) == 1:
                        meta = candidates[0]
                    else:
                        # try to disambiguate using folder ID if present
                        chosen = None
                        if id_from_folder:
                            for c in candidates:
                                cid = c.get('ID')
                                if cid is not None and str(cid).strip() == str(id_from_folder).strip():
                                    chosen = c
                                    break
                        if chosen is None:
                            # ambiguous: pick first but record ambiguity
                            meta = candidates[0]
                            candidate_ids = [c.get('ID') for c in candidates if c.get('ID') is not None]
                            print(f"Ambiguous metadata for item '{item_key}': candidate IDs = {candidate_ids}; using first entry")
                            # store ambiguity info in output later
                            ambiguous_ids = candidate_ids
                        else:
                            meta = chosen
                else:
                    meta = {}
            # meta is expected to be a dict of fields; copy over known keys to output
            entry = {
                'images': images,
            }
            # merge metadata fields if present
            if isinstance(meta, dict) and meta:
                entry.update(meta)
            else:
                # keep 'Type' key for backward compatibility if metadata was a string
                if meta:
                    entry['Type'] = meta

            # Normalize ID and Item fields in output. Prefer ID from metadata, but if missing use folder prefix.
            meta_id = None
            if isinstance(meta, dict):
                meta_id = meta.get('ID')

            if meta_id is None and id_from_folder:
                # set ID from folder (try numeric conversion)
                try:
                    meta_id = int(id_from_folder)
                except Exception:
                    meta_id = id_from_folder

            if meta_id is not None:
                entry['ID'] = meta_id

            # Ensure Item field exists and is the clean name without ID prefix
            entry['Item'] = item_key

            # Derive availability from State (case-insensitive):
            # - if State contains 'sold' or 'out' -> 'unavailable'
            # - if State contains 'in' (e.g. 'IN STOCK') -> 'available'
            state_val = entry.get('State') or entry.get('state')
            if state_val is not None:
                s = str(state_val).strip().lower()
                if 'sold' in s or 'out' in s:
                    entry['Status'] = 'Unavailable'
                elif 'in' in s:
                    entry['Status'] = 'Available'

            # If we detected ambiguity, attach candidate IDs
            if 'ambiguous_ids' in locals() and locals()['ambiguous_ids']:
                entry['ambiguous_metadata'] = locals()['ambiguous_ids']

            # Debug prints
            if meta_id is not None:
                print(f"Matched '{item_name}': Item='{item_key}', ID={meta_id}, images={len(images)}")
            else:
                print(f"Processing {item_name}: {len(images)} images found; no ID metadata")

            result['Items'][item_name] = entry

    return result


def main():
    # script location is dataSet; base_dir is this folder
    base_dir = os.path.dirname(os.path.abspath(__file__))
    metadata = read_metadata_from_excel(base_dir)
    index = gather_images(base_dir, metadata)
    # Transform the index into the requested array format
    products = []
    # index may be { 'Items': {...} } or older shapes; try to find the dict of items
    items_map = None
    if isinstance(index, dict):
        if 'Items' in index:
            items_map = index['Items']
        else:
            # try first nested dict value
            for v in index.values():
                if isinstance(v, dict):
                    items_map = v
                    break

    if not items_map:
        items_map = {}

    for folder_name, entry in items_map.items():
        prod = {}
        # id (ensure int when possible)
        pid = entry.get('ID')
        try:
            if pid is None:
                # try extract numeric prefix from folder_name
                parts = folder_name.split(' ', 1)
                if parts and parts[0].isdigit():
                    pid = int(parts[0])
            if isinstance(pid, float) and pid.is_integer():
                pid = int(pid)
            prod['id'] = pid
        except Exception:
            prod['id'] = pid

        # status from State or availability
        state_val = entry.get('State') or entry.get('state') or entry.get('availability')
        status = None
        if state_val is not None:
            s = str(state_val).strip().lower()
            if 'sold' in s or 'out' in s:
                status = 'Unavailable'
            elif 'in' in s:
                status = 'Available'
        # fallback to availability field
        if not status and entry.get('availability'):
            status = 'Available' if entry.get('availability') == 'available' else 'Unavailable'
        prod['status'] = status or 'Unknown'

        # product_name and item fields
        prod['product_name'] = entry.get('Item') or folder_name
        # size
        prod['size_eu'] = entry.get('Size (EU)') or entry.get('Size') or None
        # collection (accept both spellings)
        prod['collection'] = entry.get('Collection') or entry.get('Colection') or None
        prod['type'] = entry.get('Type') or None
        prod['price_pln'] = entry.get('Price PLN') or entry.get('Price') or None
        prod['fabric'] = entry.get('Fabric') or None
        prod['base'] = entry.get('Base') or None
        # include Website description field if present in metadata
        prod['website_des'] = entry.get('Website Des') or entry.get('WebsiteDes') or None
        prod['images'] = entry.get('images') or []

        products.append(prod)

    out_path = os.path.join(base_dir, 'images_index.json')
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(products, f, ensure_ascii=False, indent=2)
    total_images = sum(len(p.get('images') or []) for p in products)
    print(f'Wrote {out_path} with {len(products)} products and {total_images} image paths')


if __name__ == '__main__':
    main()
